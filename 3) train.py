import os
import cv2
import numpy as np
import openpyxl
from PIL import Image

class Train:
    def __init__(self):
        self.data_dir = "data"
        self.face_data = []
        self.labels = []
        self.face_recognizer = cv2.face.LBPHFaceRecognizer_create()
        self.attendance_file = "attendance.xlsx"

    def train_classifier(self):
        self.face_data, self.labels = self.load_face_data()
        if len(self.face_data) == 0:
            messagebox.showerror("Error", "No data found for training.")
            return

        self.face_recognizer.train(self.face_data, np.array(self.labels))
        self.face_recognizer.save("trained_model.xml")
        messagebox.showinfo("Success", "Classifier trained and saved successfully.")

    def load_face_data(self):
        face_data = []
        labels = []
        for subdir, dirs, files in os.walk(self.data_dir):
            for file in files:
                if file.startswith("."):
                    continue

                label = os.path.basename(subdir)
                img_path = os.path.join(subdir, file)

                pil_image = Image.open(img_path).convert("L")  # Convert to grayscale
                image_array = np.array(pil_image, "uint8")
                face = self.face_data_function(image_array)

                if face is not None:
                    face_data.append(face)
                    labels.append(int(label))

        return face_data, labels

    def face_data_function(self, image_array):
        # Implement face detection and return the face region from the image
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
        faces = face_cascade.detectMultiScale(image_array, scaleFactor=1.1, minNeighbors=5)

        for (x, y, w, h) in faces:
            face = image_array[y:y + h, x:x + w]
            return face

    def face_recognition_function(self, face):
        # Implement face recognition and return the predicted label and confidence
        label, confidence = self.face_recognizer.predict(face)
        return label, confidence

    def recognize_face(self, frame):
        gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        face_region = self.face_data_function(gray_frame)
        if face_region is not None:
            predicted_label, confidence = self.face_recognition_function(face_region)
            if confidence < 100:  # You may need to adjust the threshold based on your training data
                name = self.get_name_from_label(predicted_label)
                return name
        return None

    def get_name_from_label(self, label):
        # Replace this with your own logic to map the label to a person's name
        # Example: {"0": "Person 1", "1": "Person 2", ...}
        names = {
            0: "Person 1",
            1: "Person 2",
            # Add more names here as needed
        }

        return names.get(label, "Unknown")

    def update_attendance(self, name):
        if not os.path.isfile(self.attendance_file):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(["Name", "Attendance"])
        else:
            wb = openpyxl.load_workbook(self.attendance_file)
            sheet = wb.active

        # Check if the name already exists in the sheet
        name_column = sheet["A"]
        names = [cell.value for cell in name_column]
        if name in names:
            # Update attendance status for existing name
            row_index = names.index(name) + 1
            sheet.cell(row=row_index, column=2).value = "Present"
        else:
            # Add a new row for the new name
            new_row = [name, "Present"]
            sheet.append(new_row)

        wb.save(self.attendance_file)
        wb.close()

