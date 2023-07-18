import tkinter as tk
from tkinter import messagebox, simpledialog
from train import Train
import cv2
import os
import openpyxl
import csv

class Attendance:
    def __init__(self, attendance_file="attendance.xlsx"):
        self.attendance_file = attendance_file

    def update_attendance(self, name):
        if not os.path.isfile(self.attendance_file):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(["Name", "Attendance"])
        else:
            wb = openpyxl.load_workbook(self.attendance_file)
            sheet = wb.active

        # Check if the name already exists in the sheet
        name_column = sheet["A"]
        names = [cell.value for cell in name_column]
        if name in names:
            # Update attendance status for existing name
            row_index = names.index(name) + 1
            sheet.cell(row=row_index, column=2).value = "Present"
        else:
            # Add a new row for the new name
            new_row = [name, "Present"]
            sheet.append(new_row)

        wb.save(self.attendance_file)
        wb.close()
        
    def fetch_attendance_data(self):
        if not os.path.isfile(self.attendance_file):
            return None

        wb = openpyxl.load_workbook(self.attendance_file)
        sheet = wb.active

        attendance_data = []
        for row in sheet.iter_rows(values_only=True):
            attendance_data.append(row)

        wb.close()
        return attendance_data

    def export_to_csv(self):
        attendance_data = self.fetch_attendance_data()
        if attendance_data:
            with open("attendance.csv", mode="w", newline="") as csvfile:
                writer = csv.writer(csvfile)
                writer.writerows(attendance_data)
            print("Attendance data exported to attendance.csv")
        else:
            print("Attendance data not found.")

    def import_from_csv(self, csv_file):
        if not os.path.isfile(csv_file):
            print(f"CSV file '{csv_file}' not found.")
            return

        with open(csv_file, mode="r") as csvfile:
            reader = csv.reader(csvfile)
            attendance_data = [row for row in reader]

        if attendance_data:
            wb = openpyxl.Workbook()
            sheet = wb.active
            for row in attendance_data:
                sheet.append(row)

            wb.save(self.attendance_file)
            wb.close()
            print(f"Attendance data imported from {csv_file} to {self.attendance_file}")
        else:
            print("No data found in the CSV file.")

    def view_attendance(self):
        attendance_data = self.fetch_attendance_data()
        if attendance_data:
            for row in attendance_data:
                print(row)
        else:
            print("Attendance data not found.")

    def clear_attendance(self):
        if os.path.isfile(self.attendance_file):
            os.remove(self.attendance_file)
            print("Attendance data cleared.")
        else:
            print("Attendance data not found.")


class FaceRecognitionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Face Recognition App")
        self.train = Train()
        self.attendance = Attendance()
        self.create_widgets()

    def create_widgets(self):
        self.btn_collect_data = tk.Button(self.root, text="Collect Data", command=self.collect_data)
        self.btn_collect_data.pack(pady=10)

        self.btn_train_classifier = tk.Button(self.root, text="Train Classifier", command=self.train_classifier)
        self.btn_train_classifier.pack(pady=10)

        self.btn_recognize_face = tk.Button(self.root, text="Recognize Face", command=self.recognize_face)
        self.btn_recognize_face.pack(pady=10)

        self.btn_mark_attendance = tk.Button(self.root, text="Mark Attendance", command=self.mark_attendance, state=tk.DISABLED)
        self.btn_mark_attendance.pack(pady=10)

        self.btn_view_attendance = tk.Button(self.root, text="View Attendance", command=self.view_attendance)
        self.btn_view_attendance.pack(pady=10)

    def collect_data(self):
        # Capture images using OpenCV and save them for training
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
        video_capture = cv2.VideoCapture(0)

        person_name = simpledialog.askstring("Input", "Enter your name:")
        if not person_name:
            return

        data_path = os.path.join("data", person_name)
        os.makedirs(data_path, exist_ok=True)

        count = 0
        while count < 20:  # Capture 20 images for training
            ret, frame = video_capture.read()
            gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray_frame, scaleFactor=1.1, minNeighbors=5)

            for (x, y, w, h) in faces:
                face_region = gray_frame[y:y + h, x:x + w]
                cv2.imwrite(os.path.join(data_path, f"{person_name}_{count}.jpg"), face_region)
                count += 1

            cv2.imshow("Collecting Data", frame)
            if cv2.waitKey(1) & 0xFF == ord("q"):
                break

        video_capture.release()
        cv2.destroyAllWindows()

    def train_classifier(self):
        self.train.train_classifier()

    def recognize_face(self):
        predicted_name = self.train.recognize_face()
        if predicted_name:
            messagebox.showinfo("Face Recognized", f"Hello, {predicted_name}!")
            self.current_attendance_name = predicted_name
            self.btn_mark_attendance.config(state=tk.NORMAL)
        else:
            messagebox.showinfo("Unknown Face", "Sorry, your face is not recognized.")
            self.btn_mark_attendance.config(state=tk.DISABLED)

    def mark_attendance(self):
        if self.current_attendance_name:
            self.attendance.update_attendance(self.current_attendance_name)
            messagebox.showinfo("Attendance Marked", f"Attendance for {self.current_attendance_name} marked successfully.")
            self.btn_mark_attendance.config(state=tk.DISABLED)

    def view_attendance(self):
        self.attendance.view_attendance()


if __name__ == "__main__":
    root = tk.Tk()
    app = FaceRecognitionApp(root)
    root.mainloop()
